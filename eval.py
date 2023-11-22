# -*- coding: utf-8 -*-
"""
# @file name  : eval.py
# @author     : chenzhanpeng https://github.com/chenzpstar
# @date       : 2023-09-10
# @brief      : 图像评估
# @reference  : https://blog.csdn.net/fovever_/article/details/129332278
"""

import os
import sys

os.environ['CUDA_VISIBLE_DEVICES'] = '0'

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.join(BASE_DIR, '..'))

import cv2
import numpy as np
import torch
from natsort import natsorted
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from common import get_test_args
from core.metric import *


def eval_metrics(img1, img2, imgf):
    sd = calc_std(imgf)
    ag = calc_ag(imgf)
    sf = calc_sf(imgf)

    mse = (calc_mse(img1, imgf) + calc_mse(img2, imgf)) * 0.5
    psnr = calc_psnr(mse)

    cc = (calc_cc(img1, imgf) + calc_cc(img2, imgf)) * 0.5
    scd = calc_scd(img1, img2, imgf)

    en = calc_entropy(imgf)
    ce = calc_cross_ent(img1, imgf) + calc_cross_ent(img2, imgf)
    mi = calc_mul_info(img1, imgf, normalized=True) + calc_mul_info(
        img2, imgf, normalized=True)

    qabf, nabf, labf = calc_Qabf(img1, img2, imgf, L=1.5, full=True)

    ssim = (calc_ssim(img1, imgf) + calc_ssim(img2, imgf)) * 0.5
    msssim = (calc_msssim(img1, imgf) + calc_msssim(img2, imgf)) * 0.5

    viff = calc_viff(img1, img2, imgf, simple=False)

    return {
        'sd': sd.item(),
        'ag': ag.item(),
        'sf': sf.item(),
        'mse': mse.item(),
        'psnr': psnr.item(),
        'cc': cc.item(),
        'scd': scd.item(),
        'en': en.item(),
        'ce': ce.item(),
        'mi': mi.item(),
        'qabf': qabf.item(),
        'nabf': nabf.item(),
        'labf': labf.item(),
        'ssim': ssim.item(),
        'msssim': msssim.item(),
        'viff': viff.item(),
        # 'qabf': 1,
        # 'nabf': 1,
        # 'labf': 1,
        # 'ssim': 1,
        # 'msssim': 1,
        # 'viff': 1,
    }


def write_excel(file_name, sheet_name='test', column=0, data=None):
    try:
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        workbook = Workbook()  # 若文件不存在，则创建新文件

    # 获取或创建指定工作表
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(title=sheet_name)

    # 在指定列中插入数据
    column = get_column_letter(column + 1)
    for i, value in enumerate(data):
        cell = worksheet[column + str(i + 1)]
        cell.value = value

    # 保存文件
    workbook.save(file_name)


if __name__ == '__main__':

    import time

    torch.cuda.empty_cache()

    args = get_test_args()
    device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
    # device = torch.device('cuda:0' if args.gpu else 'cpu')
    # device = torch.device('cpu')

    sheet_name = 'method'
    # sheet_name = 'metric'

    method_list = [
        'DeepFuse', 'DenseFuse', 'VIFNet', 'DBNet', 'SEDRFuse',
        'NestFuse', 'RFNNest', 'UNFusion', 'Res2Fusion', 'MAFusion',
        'IFCNN', 'DIFNet', 'PMGI', 'PFNetv1', 'PFNetv2', 'MyFusion'
    ]
    method_names = [method_list[0]]
    exp_name = None
    # exp_name = 'exp1'

    # data_dir = os.path.join(BASE_DIR, 'data', 'samples', args.data)
    data_dir = os.path.join(BASE_DIR, '..', 'datasets', args.data)
    # save_path = os.path.join(BASE_DIR, '..', f'metrics_{args.data}.xlsx')

    if args.data in ['tno']:
        img1_dir = os.path.join(data_dir, 'vis')
        img2_dir = os.path.join(data_dir, 'ir')
    elif args.data in ['roadscene', 'msrs']:
        img1_dir = os.path.join(data_dir, 'test', 'vis')
        img2_dir = os.path.join(data_dir, 'test', 'ir')
    elif args.data in ['polar']:
        img1_dir = os.path.join(data_dir, 'test', 'vis')
        img2_dir = os.path.join(data_dir, 'test', 'po')

    if exp_name is None:
        ckpt_dir = os.path.join(BASE_DIR, '..', 'checkpoints', args.ckpt)
    else:
        ckpt_dir = os.path.join(BASE_DIR, '..', 'checkpoints', exp_name, args.ckpt)
    # save_path = os.path.join(ckpt_dir, f'metrics_{args.data}.xlsx')
    save_path = os.path.join(ckpt_dir, f'metrics_{args.data}_{method_names[0]}.xlsx')

    imgf_dir = os.path.join(ckpt_dir, args.data)

    # results_dir = os.path.join(BASE_DIR, '..', 'results', args.data)
    # save_path = os.path.join(results_dir, f'metrics_{args.data}.xlsx')

    for i, method_name in enumerate(method_names):
        sd_list = []
        ag_list = []
        sf_list = []
        mse_list = []
        psnr_list = []
        cc_list = []
        scd_list = []
        en_list = []
        ce_list = []
        mi_list = []
        qabf_list = []
        nabf_list = []
        labf_list = []
        ssim_list = []
        msssim_list = []
        viff_list = []
        name_list = []

        # 评估方法
        # imgf_dir = os.path.join(results_dir, method_name)

        print(f'evaluating {method_name} ...')
        start = time.time()

        for i, img in enumerate(natsorted(os.listdir(img1_dir))):
            # 读取数据
            img1_path = os.path.join(img1_dir, img)
            img2_path = os.path.join(img2_dir, img)
            imgf_path = os.path.join(imgf_dir, f'{i + 1:0>2}.bmp')

            img1 = cv2.imread(img1_path,
                              cv2.IMREAD_GRAYSCALE).astype(np.float32)
            img2 = cv2.imread(img2_path,
                              cv2.IMREAD_GRAYSCALE).astype(np.float32)
            imgf = cv2.imread(imgf_path,
                              cv2.IMREAD_GRAYSCALE).astype(np.float32)

            img1 = torch.from_numpy(
                img1.copy()).float().unsqueeze(0).unsqueeze(0)
            img2 = torch.from_numpy(
                img2.copy()).float().unsqueeze(0).unsqueeze(0)
            imgf = torch.from_numpy(
                imgf.copy()).float().unsqueeze(0).unsqueeze(0)

            # imgf = torch.rand_like(img1)

            img1.to(device, non_blocking=True)
            img2.to(device, non_blocking=True)
            imgf.to(device, non_blocking=True)

            # 评估图像
            print(f'evaluating {img} ...')

            with torch.no_grad():
                results = eval_metrics(img1, img2, imgf)

            # 记录结果
            sd_list.append(results['sd'])
            ag_list.append(results['ag'])
            sf_list.append(results['sf'])
            mse_list.append(results['mse'])
            psnr_list.append(results['psnr'])
            cc_list.append(results['cc'])
            scd_list.append(results['scd'])
            en_list.append(results['en'])
            ce_list.append(results['ce'])
            mi_list.append(results['mi'])
            qabf_list.append(results['qabf'])
            nabf_list.append(results['nabf'])
            labf_list.append(results['labf'])
            ssim_list.append(results['ssim'])
            msssim_list.append(results['msssim'])
            viff_list.append(results['viff'])
            name_list.append(img)

        end = time.time()
        print(f'evaluating {method_name} done, cost {end - start:.3f}s')

        # 计算均值
        sd_list.insert(0, np.mean(sd_list))
        ag_list.insert(0, np.mean(ag_list))
        sf_list.insert(0, np.mean(sf_list))
        mse_list.insert(0, np.mean(mse_list))
        psnr_list.insert(0, np.mean(psnr_list))
        cc_list.insert(0, np.mean(cc_list))
        scd_list.insert(0, np.mean(scd_list))
        en_list.insert(0, np.mean(en_list))
        ce_list.insert(0, np.mean(ce_list))
        mi_list.insert(0, np.mean(mi_list))
        qabf_list.insert(0, np.mean(qabf_list))
        nabf_list.insert(0, np.mean(nabf_list))
        labf_list.insert(0, np.mean(labf_list))
        ssim_list.insert(0, np.mean(ssim_list))
        msssim_list.insert(0, np.mean(msssim_list))
        viff_list.insert(0, np.mean(viff_list))
        name_list.insert(0, 'mean')

        # 计算标准差
        sd_list.insert(1, np.std(sd_list))
        ag_list.insert(1, np.std(ag_list))
        sf_list.insert(1, np.std(sf_list))
        mse_list.insert(1, np.std(mse_list))
        psnr_list.insert(1, np.std(psnr_list))
        cc_list.insert(1, np.std(cc_list))
        scd_list.insert(1, np.std(scd_list))
        en_list.insert(1, np.std(en_list))
        ce_list.insert(1, np.std(ce_list))
        mi_list.insert(1, np.std(mi_list))
        qabf_list.insert(1, np.std(qabf_list))
        nabf_list.insert(1, np.std(nabf_list))
        labf_list.insert(1, np.std(labf_list))
        ssim_list.insert(1, np.std(ssim_list))
        msssim_list.insert(1, np.std(msssim_list))
        viff_list.insert(1, np.std(viff_list))
        name_list.insert(1, 'std')

        if sheet_name == 'method':
            # 插入名字
            sd_list.insert(0, f'{"SD"}')
            ag_list.insert(0, f'{"AG"}')
            sf_list.insert(0, f'{"SF"}')
            mse_list.insert(0, f'{"MSE"}')
            psnr_list.insert(0, f'{"PSNR"}')
            cc_list.insert(0, f'{"CC"}')
            scd_list.insert(0, f'{"SCD"}')
            en_list.insert(0, f'{"EN"}')
            ce_list.insert(0, f'{"CE"}')
            mi_list.insert(0, f'{"MI"}')
            qabf_list.insert(0, f'{"Qabf"}')
            nabf_list.insert(0, f'{"Nabf"}')
            labf_list.insert(0, f'{"Labf"}')
            ssim_list.insert(0, f'{"SSIM"}')
            msssim_list.insert(0, f'{"MSSSIM"}')
            viff_list.insert(0, f'{"VIFF"}')
            name_list.insert(0, '')

            # 写入文件
            write_excel(save_path, method_name, 0, name_list)
            write_excel(save_path, method_name, 1, sd_list)
            write_excel(save_path, method_name, 2, ag_list)
            write_excel(save_path, method_name, 3, sf_list)
            write_excel(save_path, method_name, 4, mse_list)
            write_excel(save_path, method_name, 5, psnr_list)
            write_excel(save_path, method_name, 6, cc_list)
            write_excel(save_path, method_name, 7, scd_list)
            write_excel(save_path, method_name, 8, en_list)
            write_excel(save_path, method_name, 9, ce_list)
            write_excel(save_path, method_name, 10, mi_list)
            write_excel(save_path, method_name, 11, qabf_list)
            write_excel(save_path, method_name, 12, nabf_list)
            write_excel(save_path, method_name, 13, labf_list)
            write_excel(save_path, method_name, 14, ssim_list)
            write_excel(save_path, method_name, 15, msssim_list)
            write_excel(save_path, method_name, 16, viff_list)

        elif sheet_name == 'metric':
            # 插入名字
            sd_list.insert(0, f'{method_name}')
            ag_list.insert(0, f'{method_name}')
            sf_list.insert(0, f'{method_name}')
            mse_list.insert(0, f'{method_name}')
            psnr_list.insert(0, f'{method_name}')
            cc_list.insert(0, f'{method_name}')
            scd_list.insert(0, f'{method_name}')
            en_list.insert(0, f'{method_name}')
            ce_list.insert(0, f'{method_name}')
            mi_list.insert(0, f'{method_name}')
            qabf_list.insert(0, f'{method_name}')
            nabf_list.insert(0, f'{method_name}')
            labf_list.insert(0, f'{method_name}')
            ssim_list.insert(0, f'{method_name}')
            msssim_list.insert(0, f'{method_name}')
            viff_list.insert(0, f'{method_name}')
            name_list.insert(0, '')

            # 写入文件
            if i == 0:
                write_excel(save_path, 'SD', 0, name_list)
                write_excel(save_path, 'AG', 0, name_list)
                write_excel(save_path, 'SF', 0, name_list)
                write_excel(save_path, 'MSE', 0, name_list)
                write_excel(save_path, 'PSNR', 0, name_list)
                write_excel(save_path, 'CC', 0, name_list)
                write_excel(save_path, 'SCD', 0, name_list)
                write_excel(save_path, 'EN', 0, name_list)
                write_excel(save_path, 'CE', 0, name_list)
                write_excel(save_path, 'MI', 0, name_list)
                write_excel(save_path, 'Qabf', 0, name_list)
                write_excel(save_path, 'Nabf', 0, name_list)
                write_excel(save_path, 'Labf', 0, name_list)
                write_excel(save_path, 'SSIM', 0, name_list)
                write_excel(save_path, 'MSSSIM', 0, name_list)
                write_excel(save_path, 'VIFF', 0, name_list)

            write_excel(save_path, 'SD', i + 1, sd_list)
            write_excel(save_path, 'AG', i + 1, ag_list)
            write_excel(save_path, 'SF', i + 1, sf_list)
            write_excel(save_path, 'MSE', i + 1, mse_list)
            write_excel(save_path, 'PSNR', i + 1, psnr_list)
            write_excel(save_path, 'CC', i + 1, cc_list)
            write_excel(save_path, 'SCD', i + 1, scd_list)
            write_excel(save_path, 'EN', i + 1, en_list)
            write_excel(save_path, 'CE', i + 1, ce_list)
            write_excel(save_path, 'MI', i + 1, mi_list)
            write_excel(save_path, 'Qabf', i + 1, qabf_list)
            write_excel(save_path, 'Nabf', i + 1, nabf_list)
            write_excel(save_path, 'Labf', i + 1, labf_list)
            write_excel(save_path, 'SSIM', i + 1, ssim_list)
            write_excel(save_path, 'MSSSIM', i + 1, msssim_list)
            write_excel(save_path, 'VIFF', i + 1, viff_list)
